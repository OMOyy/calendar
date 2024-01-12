import os
import openpyxl
import json
from datetime import datetime
wb = openpyxl.load_workbook('test.xlsx', data_only=True) 
s1 = wb['工作表1']

#[]陣列
#{}物件
def convert_to_json_serializable(obj):
    if isinstance(obj, datetime):
        return obj.strftime('%Y-%m-%d %H:%M:%S')
    raise TypeError("Type not serializable")

def get_values(sheet):
    # 获取表格的列数
    num_columns = sheet.max_column

    # 获取属性名，即第一行的数据
    headers = [sheet.cell(row=1, column=col).value for col in range(1, num_columns + 1)]

    # 存储所有对象的列表
    data = []

    # 从第二行开始遍历数据行
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 创建一个空字典，用于存储对象的属性和内容
        obj = {}
        for col, value in enumerate(row):
            # 使用属性名作为键，将值存入字典
            obj[headers[col]] = value
        # 将对象添加到列表中
        data.append(obj)

    return data

print(get_values(s1)) 

# 使用函数获取数据并创建对象
my_object = get_values(s1)
# 将对象写入 JSON 文件
json_file = 'output.json'
with open(json_file, 'w', encoding='utf-8') as f:
    json.dump(my_object, f, ensure_ascii=False, default=convert_to_json_serializable, indent=2)

print(f"Objects have been exported to '{json_file}'.")
